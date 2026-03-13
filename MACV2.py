import paramiko
from scp import SCPClient
import pandas as pd
import traceback
import re
import os
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from cryptography.fernet import Fernet
from dotenv import dotenv_values
import io
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.client_credential import ClientCredential
import time

class ejecutor:
    
    try:    

        def __init__(self):
            
            self.script_path = r"C:\Users\PC\OneDrive - INDEPENDENCE S.A\GerenciaTI - Documentos 1\General\2.Operación\0. Telco\1. CONECTIVIDAD\1. CONECTIVIDAD STARLINK\1. Automatizacion Macs\MAC-1\MACV2.py"
            self.smtp_user,self.smtp_password, self.ssh_key2,self.site_url,self.client_id,self.client_secret = None,None,None,None,None,None
            self.smtp_user,self.smtp_password, self.ssh_key2,self.site_url,self.client_id,self.client_secret=self.contraseñas(None,None,None,None,None,None)
            self.credentials = ClientCredential(self.client_id, self.client_secret)
            self.ctx = ClientContext(self.site_url).with_credentials(self.credentials)
            self.last_script_run = time.time()

        def contraseñas(self,smtp_user,smtp_password,ssh_key2,site_url,client_id,client_secret):
            # 1. Cargar la clave
            with open(r"C:\Users\PC\OneDrive - INDEPENDENCE S.A\GerenciaTI - Documentos 1\General\2.Operación\0. Telco\1. CONECTIVIDAD\1. CONECTIVIDAD STARLINK\1. Automatizacion Macs\variables\clave.key", "rb") as key_file:
                key = key_file.read()
            fernet = Fernet(key)
            
            # 2. Leer y desencriptar el archivo
            with open(r"C:\Users\PC\OneDrive - INDEPENDENCE S.A\GerenciaTI - Documentos 1\General\2.Operación\0. Telco\1. CONECTIVIDAD\1. CONECTIVIDAD STARLINK\1. Automatizacion Macs\variables\.env.enc", "rb") as enc_file:
                encrypted_data = enc_file.read()
            decrypted_data = fernet.decrypt(encrypted_data).decode()

            # ✅ 3. Pasar el texto desencriptado como archivo virtual
            env_vars = dotenv_values(stream=io.StringIO(decrypted_data))
            
            # 4. Usar las variables sin imprimirlas
            smtp_user = env_vars.get("SMTP_USER")
            smtp_password = env_vars.get("SMTP_PASS")
            ssh_key2 = env_vars.get("ssh_key")
            site_url = env_vars.get("site_url")
            client_id = env_vars.get("client_id")
            client_secret = env_vars.get("client_secret")

            return smtp_user,smtp_password,ssh_key2,site_url,client_id,client_secret            
        def sharepoint_online(self):

            lista = self.ctx.web.lists.get_by_title("Aprobacion")
            items = lista.items.get().top(100).execute_query()
            if len(items) == 0:
                
                return pd.DataFrame()  # Retorna un DataFrame vacío si no hay items
            else:
                data = [item.properties for item in items]
                df = pd.DataFrame(data)
                df = df[["FECHA","AV","MAC", "TIPO","PROPIEDAD","identificacion","NOMBRE","CARGO","TORRE","CORREO","ESTADO"]]
               
                return df   
        def sharepoint(self, Archivo, Hoja, Tabla_aprobaciones, estado):

            # 📁 Ruta relativa del archivo desde la raíz del sitio
            relative_path = f"/sites/ActividadesGerenciaTI/Documentos compartidos/General/2.Operación/0. Telco/1. CONECTIVIDAD/1. CONECTIVIDAD STARLINK/1. Automatizacion Macs/MAC-1/{Archivo}"

            # 📥 Descargar archivo a local
            file = self.ctx.web.get_file_by_server_relative_url(relative_path)

            if estado == "descargar":
                with open(Archivo, "wb") as local_file:
                    file.download(local_file).execute_query()

                df_macs = pd.read_excel(Archivo, sheet_name=Hoja)
                Tabla_aprobaciones = df_macs.dropna(subset=["MAC", "NOMBRE", "TORRE"]).tail(len(df_macs))
                return Tabla_aprobaciones
            
            if estado == "subir":
                
                with open(Archivo, "rb") as f:
                    self.ctx.web.get_folder_by_server_relative_url(
                        "/sites/ActividadesGerenciaTI/Documentos compartidos/General/2.Operación/0. Telco/1. CONECTIVIDAD/1. CONECTIVIDAD STARLINK/1. Automatizacion Macs/MAC-1"
                    ).upload_file(Archivo, f).execute_query()
                #print("✅ Archivo actualizado en SharePoint")
                return None                          
        def ejecutar_clase_directamente(self,Inclusion):
            try:
                if Inclusion=="Aprobados":
                    print("🚀 Ejecutando clase inclusion_mac directamente...")
                    proceso = inclusion_mac()
                    proceso.main()
                    print("✅ Clase Inclusion ejecutada exitosamente")
                    return True
                if Inclusion=="Fallos":
                    print("🚀 Ejecutando clase fallos...")
                    proceso = Inclusion_fallos()
                    proceso.main()
                    print("✅ Clase Fallos ejecutada exitosamente")         
            except Exception as e:
                print(f"❌ Error ejecutando clase: {e}")
                return False  
        def modificacion(self):
            lista = self.ctx.web.lists.get_by_title("Aprobacion")
            self.ctx.load(lista, ["LastItemModifiedDate"])
            self.ctx.execute_query()
            Actual_modificacion=lista.properties["LastItemModifiedDate"]
            return Actual_modificacion
        def main(self):
                
                self.smtp_user,self.smtp_password, self.ssh_key2,self.site_url,self.client_id,self.client_secret=self.contraseñas(None,None,None,None,None,None)
                self.credentials = ClientCredential(self.client_id, self.client_secret)
                self.ctx = ClientContext(self.site_url).with_credentials(self.credentials) 
            
                try:  
                        if self.sharepoint_online().empty:
                            print("vacio")
                        else:
                            self.ejecutar_clase_directamente("Aprobados")
                    

                except Exception as e:
                            print(f"⚠️ Error en ejecutor: {e}")
                            error_detalle = traceback.format_exc()
                            print(f"⚠️ Error en ejecutor: {e}",error_detalle)
     
    except Exception as e:
            if "503" in str(e):
                print(f"⚠️ SharePoint no disponible (503). Reintentando en 10 segundos... (Intento 1)")
                time.sleep(10)

            else: 
                print("error en el ejecutor")
        
class inclusion_mac(ejecutor):
     
    try:
        def __init__(self):
            super().__init__()                         
        def modificar_columna(self,mac):
            lista = self.ctx.web.lists.get_by_title("Aprobacion")
            items = lista.items.get().top(100).execute_query()
            for item in items:
                if item.properties["MAC"] == mac:
                    
                    item.set_property("ESTADO", "PROCESADO")
                    item.update()
            self.ctx.execute_query() 
        def Tabla_Aprobado(self, df_resultado):
            df3=pd.DataFrame(columns=["FECHA","AV","MAC", "TIPO","PROPIEDAD","identificacion","NOMBRE","CARGO","TORRE","CORREO","ESTADO","IP"])
            lista=[]
            # === Obtener el último registro del libro de MACs ===
            for _, fila in df_resultado.iterrows():

                torre_objetivo = fila["TORRE"]
                mac_objetivo = fila["MAC"]
                self.modificar_columna(mac_objetivo)
                    # Filtra el data frame por la torre objetivo
                df_filtrado = self.df_listado[self.df_listado["TORRE"].astype(str).str.strip() == str(torre_objetivo).strip()].copy()
                
                # Crea un data frame con estas variables
                df_filtrado["MAC"] = mac_objetivo
                df_filtrado["NOMBRE"] = fila["NOMBRE"]
                df_filtrado["ESTADO"] = "PENDIENTE"  # Estado 
                df_filtrado["TIPO"] = fila["TIPO"]
                df_filtrado["CORREO"] = fila["CORREO"]
                df_filtrado["PROPIEDAD"] = fila["PROPIEDAD"]
                df_filtrado["FECHA"] = fila["FECHA"]
                df_filtrado["identificacion"] = fila["identificacion"]
                df_filtrado["CARGO"] = fila["CARGO"]
                df_filtrado["AV"] = fila["AV"]
                lista.append(df_filtrado)
                
            
            if not lista:
                print("No hay filas para procesar (torres no coinciden)")
                return
            
            df3 = (pd.concat(lista, ignore_index=True).reindex(columns=df3.columns).sort_values(by=[
              "FECHA","AV","MAC","TIPO","PROPIEDAD","identificacion","NOMBRE","CARGO","TORRE","CORREO","ESTADO","IP"]).reset_index(drop=True))
            
            #df3 es el dataframe ya con los ip con su mac      
            if df3.empty:
                print("Listado de Sharepoint vacio")
                return
            else: 
                print("si hay mac por incluir")    
                self.recorrer(df3)
        def buscar_coincidencias(self,nombre_nuevo, nombres_existentes,dispositivo_especial):
            
            nombre_nuevo = nombre_nuevo
            coincidencias = []
            if dispositivo_especial:
                if nombre_nuevo in nombres_existentes:
                    coincidencias.append(nombre_nuevo)

            else:
                nombre_nuevo = nombre_nuevo.lower().split()
                for nombre_base in nombres_existentes:
                    
                    palabras_base = nombre_base.lower().split()
                    coincidencias_palabras = sum(1 for palabra in palabras_base if palabra in nombre_nuevo)

                if coincidencias_palabras >= 4:  # mínimo 2 palabras coinciden
                    coincidencias.append(nombre_base)
            
            return coincidencias
        def recorrer(self, df_resultado):
            #Agrupo por ip para que se incluyan las mac por torres
            dispositivos = df_resultado.groupby('IP')
            self.procesar_dispositivo(df_resultado,dispositivos)

            # Agrupa aprobados-replicados, y errores-pendientes
            filtro = df_resultado[df_resultado["ESTADO"].isin(["APROBADO", "REPLICADO"])]
            grupos = filtro.groupby(["TORRE", "TIPO", "NOMBRE", "MAC", "CORREO","FECHA","PROPIEDAD"]).size().reset_index()

            # Se envia los correos para los aprobados
            for _, fila in grupos.iterrows():
                print("entro a correos")
                torre = fila["TORRE"]
                torre2 = torre.replace(" ", "-")
                contraseña = torre2.replace("-", "")
                nombre = fila["NOMBRE"]
                mac = fila["MAC"]
                correo2 = fila["CORREO"]
                tipo = fila["TIPO"]
                propiedad=fila["PROPIEDAD"]
                self.correo(mac, torre2, nombre, correo2, tipo, contraseña,propiedad)

        # Se insertan los resultados a la tabla aprobados 
            self.insertar_tabla("Aprobados.xlsx", "Sheet2", df_resultado)            
        def procesar_dispositivo(self, df_resultado, dispositivos):
            """
            Procesa múltiples dispositivos (IPs), descargando su configuración,
            agregando nuevas MACs y realizando los cambios necesarios.
            
            CAMBIOS PRINCIPALES:
            1. Se procesa TODA la lista de filas por dispositivo antes de escribir/ejecutar
            2. Se escribe el archivo UNA SOLA VEZ por dispositivo
            3. Se ejecutan comandos SSH UNA SOLA VEZ por dispositivo
            4. El return está FUERA del loop de filas
            """
            
            for ip, grupo in dispositivos:
                
                usuario = "ubnt"
                ruta_remota = "/tmp/system.cfg"
                carpeta_temporal = f"C:/Users/PC/temp_cfgs/"
                Aprobados = pd.read_excel(r"C:\Users\PC\OneDrive - INDEPENDENCE S.A\GerenciaTI - Documentos 1\General\2.Operación\0. Telco\1. CONECTIVIDAD\1. CONECTIVIDAD STARLINK\1. Automatizacion Macs\MAC-1\Aprobados.xlsx", sheet_name="Sheet2")
                
                os.makedirs(carpeta_temporal, exist_ok=True)
                ruta_local = os.path.join(carpeta_temporal, f"system_{ip.replace('.', '_')}.cfg")
                
                ssh = None
                try:
                    # Credenciales de acceso
                    clave = paramiko.RSAKey.from_private_key_file(self.ssh_key2)
                    ssh = paramiko.SSHClient()
                    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    ssh.connect(
                        ip,
                        username=usuario,
                        pkey=clave,
                        disabled_algorithms=dict(pubkeys=["rsa-sha2-256", "rsa-sha2-512"]),
                        timeout=10
                    )
                    print(f"✅ Conexión establecida con {ip}")

                    # Descargar archivo de configuración
                    with SCPClient(ssh.get_transport()) as scp:
                        scp.get(ruta_remota, ruta_local)
                    
                    with open(ruta_local, "r", encoding="utf-8") as f:
                        contenido = f.read()

                    # Extraer información del archivo de configuración
                    indices = [int(m.group(1)) for m in re.finditer(r"wireless\.1\.mac_acl\.(\d+)\.mac=", contenido)]
                    siguiente = max(indices, default=-1) + 1
                    
                    # Crear conjuntos con MACs y nombres existentes
                    mac_existentes = set(m.group(1).lower() for m in re.finditer(r"mac_acl\.\d+\.mac=(.*?)\n", contenido))
                    nombres_existentes = set(m.group(1) for m in re.finditer(rf"mac_acl\.\d+\.comment=(.*?)\n", contenido))
                    
                    
                    nuevas_macs = 0
                    cambios_realizados = False

                    # ========== PROCESAR TODAS LAS FILAS DEL DISPOSITIVO ==========
                    for idx, fila in grupo.iterrows():
                        mac = fila["MAC"].lower()
                        av = fila["AV"]
                        propiedad = fila["PROPIEDAD"]
                        tipo = fila["TIPO"]
                        cedula = fila["identificacion"]
                        nombre_base = str(fila["NOMBRE"]).strip()
                        dispositivo_especial = False
                        print(f"Procesando fila para IP {ip}: MAC={mac}, NOMBRE={nombre_base}")
                        # Construir nombre según reglas
                        nombre = tipo[:3] + " " + propiedad[:1] + " " + nombre_base
                        
                        # Buscar nombre en aprobados
                        cedula_aprobada = Aprobados.loc[Aprobados["identificacion"] == cedula, "NOMBRE"]
                        if not cedula_aprobada.empty:
                            
                            nombre = tipo[:3] + " " + propiedad[:1] + " " + cedula_aprobada.iloc[0]
                            print("cedula encontrada en aprobados", nombre)
                        # Verificar si AV tiene valor
                        if av != "None" and not pd.isna(av) and str(av).strip() != "":
                            
                            nombre = av
                            dispositivo_especial = True
                            print("AV tiene valor",nombre)
                        # Dispositivos especiales
                        if tipo == "BIOMÉTRICO" or tipo == "IMPRESORA":
                            nombre = tipo
                            dispositivo_especial = True
                        # Verificar si MAC ya existe
                        if mac in mac_existentes:
                            print(f"  [{ip}] MAC ya existe, se omite: {mac}")
                            df_resultado.loc[idx, "ESTADO"] = "REPLICADO"
                            continue

                        # Buscar coincidencias con nombres existentes
                        coincidencias = self.buscar_coincidencias(nombre, nombres_existentes,dispositivo_especial)
                        
                        if coincidencias:
                            # Caso: Nombre encontrado en el listado, reemplazar MAC
                            print("Nombre se encuentra en el listado")
                            nombre_coincidencia = coincidencias[0]
                            numero = re.search(rf"mac_acl\.(\d+)\.comment={nombre_coincidencia}\n", contenido).group(1)
                            mac_anterior = re.search(rf"mac_acl\.{numero}\.mac=(.*?)\n", contenido).group(1)
                            contenido = re.sub(rf"mac_acl\.{numero}\.mac={mac_anterior}\n", rf"mac_acl.{numero}.mac={mac}\n", contenido)
                            
                            df_resultado.loc[idx, "ESTADO"] = "REPLICADO"
                            nuevas_macs += 1
                            cambios_realizados = True
                        else:
                            # Caso: Nombre no encontrado, agregar nueva entrada
                            bloque = f"""
wireless.1.mac_acl.{siguiente}.mac={mac}
wireless.1.mac_acl.{siguiente}.status=enabled
wireless.1.mac_acl.{siguiente}.comment={nombre}"""
                            
                            contenido += bloque
                            siguiente += 1
                            nuevas_macs += 1
                            cambios_realizados = True
                            print(f"  [{ip}] MAC agregada: {mac} - {nombre}")
                            df_resultado.loc[idx, "ESTADO"] = "APROBADO"

                    # ========== ESCRIBIR ARCHIVO UNA SOLA VEZ ==========
                    if cambios_realizados:
                        with open(ruta_local, "w", encoding="utf-8") as f:
                            f.write(contenido)
                        
                        # Subir archivo modificado
                        with SCPClient(ssh.get_transport()) as scp:
                            scp.put(ruta_local, ruta_remota)
                        
                        # ========== EJECUTAR COMANDOS SSH UNA SOLA VEZ ==========
                        comandos = [
                            "lock /var/lock/.system.cfg.lock",
                            "cp /tmp/system.cfg /etc/system.cfg",
                            "cfgmtd -f /etc/system.cfg -w",
                            "wlanconfig ath0 listmac",
                            "iwpriv ath0 maccmd 1",
                            "iwpriv ath0 maccmd 4",
                            "lock -u /var/lock/.system.cfg.lock"
                        ]
                        for cmd in comandos:
                            stdin, stdout, stderr = ssh.exec_command(cmd)
                            stdout.read().decode()
                            stderr.read().decode()
                        
                        print(f"✅ [{ip}] {nuevas_macs} MACs procesadas correctamente")
                    else:
                        print(f"ℹ️ [{ip}] No se realizaron cambios")

                except Exception as e:
                    print(f"❌ [{ip}] Error: {str(e)}")
                    
                    for idx, fila in grupo.iterrows():
                        df_resultado.loc[idx, "ESTADO"] = "ERROR"

                finally:
                    if ssh is not None:
                        ssh.close()
                    if os.path.exists(ruta_local):
                        os.remove(ruta_local)

        def eliminar_macs_dispositivo(self, ips, macs):
            """Elimina MACs de la configuración de los dispositivos y aplica cambios.

            Args:
                ips (list): Lista de direcciones IP de los equipos.
                macs (iterable): Lista de MACs (cadenas) a eliminar en cada dispositivo.

            Retorna:
                dict: Diccionario con IPs como claves y bool como valores (True si se aplicaron cambios).
            """
            macs_a_eliminar = {m.lower().strip() for m in macs if m}
            if not macs_a_eliminar:
                return {ip: False for ip in ips}

            resultados = {}
            for ip in ips:
                resultados[ip] = self._eliminar_macs_en_ip(ip, macs_a_eliminar)
            return resultados

        def _eliminar_macs_en_ip(self, ip, macs_a_eliminar):
            usuario = "ubnt"
            ruta_remota = "/tmp/system.cfg"
            carpeta_temporal = f"C:/Users/PC/temp_cfgs/"
            os.makedirs(carpeta_temporal, exist_ok=True)
            ruta_local = os.path.join(carpeta_temporal, f"system_{ip.replace('.', '_')}.cfg")

            ssh = None
            try:
                clave = paramiko.RSAKey.from_private_key_file(self.ssh_key2)
                ssh = paramiko.SSHClient()
                ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                ssh.connect(
                    ip,
                    username=usuario,
                    pkey=clave,
                    disabled_algorithms=dict(pubkeys=["rsa-sha2-256", "rsa-sha2-512"]),
                    timeout=10
                )
                print(f"✅ Conexión establecida con {ip} (eliminar MACs)")

                with SCPClient(ssh.get_transport()) as scp:
                    scp.get(ruta_remota, ruta_local)

                with open(ruta_local, "r", encoding="utf-8") as f:
                    contenido = f.read()

                original = contenido

                for m in re.finditer(r"wireless\.1\.mac_acl\.(\d+)\.mac=(.*?)\n", contenido, flags=re.IGNORECASE):
                    idx = m.group(1)
                    mac_actual = m.group(2).strip().lower()
                    if mac_actual in macs_a_eliminar:
                        # Eliminar mac + status + comment para ese índice
                        for suffix in ["mac", "status", "comment"]:
                            contenido = re.sub(
                                rf"wireless\.1\.mac_acl\.{idx}\.{suffix}=.*\n",
                                "",
                                contenido,
                                flags=re.IGNORECASE,
                            )

                # Normalizar saltos de línea consecutivos
                contenido = re.sub(r"\n{3,}", "\n\n", contenido)

                if contenido == original:
                    print(f"ℹ️ [{ip}] No se detectaron MACs para eliminar")
                    return False

                with open(ruta_local, "w", encoding="utf-8") as f:
                    f.write(contenido)

                with SCPClient(ssh.get_transport()) as scp:
                    scp.put(ruta_local, ruta_remota)

                comandos = [
                    "lock /var/lock/.system.cfg.lock",
                    "cp /tmp/system.cfg /etc/system.cfg",
                    "cfgmtd -f /etc/system.cfg -w",
                    "wlanconfig ath0 listmac",
                    "iwpriv ath0 maccmd 1",
                    "iwpriv ath0 maccmd 4",
                    "lock -u /var/lock/.system.cfg.lock"
                ]
                for cmd in comandos:
                    stdin, stdout, stderr = ssh.exec_command(cmd)
                    stdout.read().decode()
                    stderr.read().decode()

                print(f"✅ [{ip}] MACs eliminadas: {', '.join(sorted(macs_a_eliminar))}")
                return True

            except Exception as e:
                print(f"❌ [{ip}] Error eliminando MACs: {e}")
                return False

            finally:
                if ssh is not None:
                    ssh.close()
                if os.path.exists(ruta_local):
                    os.remove(ruta_local)

        def insertar_tabla(self, tabla, N_hoja, df_resultado):
            
            # Cargar libro y buscar última fila usada
            wb = load_workbook(tabla)
            hoja = wb[N_hoja]
            fila_inicio = hoja.max_row  # última fila con datos
            wb.close()
            with pd.ExcelWriter(tabla, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df_resultado.to_excel(writer, sheet_name=N_hoja, index=False, header=False, startrow=fila_inicio)
            self.sharepoint(tabla, N_hoja, "", "subir")            
        def borrado_aprobaciones(self):
            lista = self.ctx.web.lists.get_by_title("Aprobacion")
            items = lista.items.get().top(100).execute_query()
            for item in items:
                if item.properties["ESTADO"] == "PROCESADO":
                    item.delete_object()
                    self.ctx.execute_query()
            
            #print("✅ Filas eliminadas, se conserva la primera fila como encabezado")              
        def borrado_duplicados(self, tabla, N_hoja):
            self.sharepoint(tabla, N_hoja, "", "descargar")
            wb = load_workbook(tabla)
            ws = wb[N_hoja]    
            vistos = set()
            filas_a_borrar = []

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Convertimos cada valor de la fila a string para comparación
                fila_como_texto = tuple(str(x) if x is not None else "" for x in row)

                if fila_como_texto in vistos:
                    filas_a_borrar.append(idx)
                else:
                    vistos.add(fila_como_texto)

            # Borrar de abajo hacia arriba
            for idx in reversed(filas_a_borrar):
                ws.delete_rows(idx)
            
            #print("✅ Duplicados eliminados")
            wb.save(tabla)
            self.sharepoint(tabla, N_hoja, "", "subir")
            wb.close()       
        def correo(self, MAC, torre, nombre, dominio, tipo, contraseña,propiedad):
            msg = MIMEMultipart("related")
            msg["Subject"] = f"INCLUSION DE MAC - {torre}"
            msg["From"] = "starlink@independence.com.co"
            msg["Reply-To"] = "mesadeayuda@independence.com.co"  
            msg["To"] = dominio
            msg["Cc"] ="cfgarzon@independence.com.co,mesadeayuda@independence.com.co,soporteticampo@independence.com.co"
            
            html = f"""\
    <html>
    <body>
        <p><b>Buen día,</b><br><br>
        Se confirma inclusión de la MAC solicitada:<br><br><br>

        <b>Nombre:</b> {nombre} <br>
        <b>MAC:</b> {MAC} <br>
        <b>Dispositivo:</b> {tipo} <br>
        <b>Dispositivo:</b> {propiedad} <br>
        <b>Torre:</b> {torre} <br><br>
            
        a continuación, relacionamos la información para establecer conexión: <br><br><br>

        <b>Nombre de la red :</b> IND{torre} <br>
        <b>Contraseña:</b> IND{contraseña} <br><br><br>
        

        <u><i>No olvide desactivar la opción de <b>MAC aleatoria</b> en el dispositivo para poder establecer conexión.</i></u><br><br>
        En caso de no lograr conexion siga lo siguientes pasos:</b> 
        <br><br>
        <div>
    <img src="cid:WINDOWS" style="display:inline; width:210px; height:auto; margin-right:10px;">
    <img src="cid:ANDROID" style="display:inline; width:200px; height:auto; margin-right:10px;">
    <img src="cid:IOS" style="display:inline; width:200px; height:auto;margin-right:10px;">
    </div>
        </p>
    </body>
    </html>
    """
            parte_html = MIMEText(html, "html")
            msg.attach(parte_html)
            
            ruta_imagen = {
                "WINDOWS": "C:/Users/PC/INDEPENDENCE S.A/GerenciaTI - 0. Telco/1. CONECTIVIDAD/1. CONECTIVIDAD STARLINK/Automatizacion Macs/MAC Windows.JPG",
                "ANDROID": "C:/Users/PC/INDEPENDENCE S.A/GerenciaTI - 0. Telco/1. CONECTIVIDAD/1. CONECTIVIDAD STARLINK/Automatizacion Macs/MAC ANDROID.JPG",
                "IOS": "C:/Users/PC/INDEPENDENCE S.A/GerenciaTI - 0. Telco/1. CONECTIVIDAD/1. CONECTIVIDAD STARLINK/Automatizacion Macs/MAC iOS.JPG"
            }
            
            for cid, ruta in ruta_imagen.items():
                with open(ruta, "rb") as f:
                    img = MIMEImage(f.read(), _subtype="jpeg")
                    img.add_header("Content-ID", f"<{cid}>")
                    img.add_header("Content-Disposition", "inline", filename=ruta.split("\\")[-1])
                    msg.attach(img)

            with smtplib.SMTP("smtp.office365.com", 587) as server:
                server.starttls()
                server.login(self.smtp_user, self.smtp_password)
                server.send_message(msg)     
        def main(self):
    
            #conexion a el shareponit
            self.smtp_user,self.smtp_password, self.ssh_key2,self.site_url,self.client_id,self.client_secret=self.contraseñas(None,None,None,None,None,None)
            self.credentials = ClientCredential(self.client_id, self.client_secret)
            self.ctx = ClientContext(self.site_url).with_credentials(self.credentials)
            self.df_listado = pd.read_excel(r"C:\Users\PC\OneDrive - INDEPENDENCE S.A\GerenciaTI - Documentos 1\General\2.Operación\0. Telco\1. CONECTIVIDAD\1. CONECTIVIDAD STARLINK\1. Automatizacion Macs\MAC-1\Listado.xlsx")
            
            # Obtener datos de sharepoint
            df_resultado=self.sharepoint_online()
           
            # Procesar tabla
            self.Tabla_Aprobado(df_resultado)
            self.borrado_aprobaciones()
            self.borrado_duplicados("Aprobados.xlsx", "Sheet2")
            
            print("\n📊 Proceso completado. Resultados guardados")
    except Exception as e:
            # Capturar stacktrace completo
            error_detalle = traceback.format_exc()

            # Construir correo
            msg = MIMEMultipart("related")
            msg["Subject"] = "⚠️ FALLA EN FLUJO"
            msg["From"] = "starlink@independence.com.co"
            msg["To"] = "cfgarzon@independence.com.co"

            html = f"""
<html>
<body>
    <p><b>Buen día,</b><br><br>
    Se confirma <b>FALLA EN FLUJO</b>.<br><br>
    <b>Error:</b> {e}<br><br>
    <pre>{error_detalle}</pre>
</body>
</html>
"""
            msg.attach(MIMEText(html, "html"))

            # Enviar correo
            with smtplib.SMTP("smtp.office365.com", 587) as server:
                server.starttls()
                server.login(self.smtp_user, self.smtp_password)
                server.send_message(msg)
                print("✅ Correo enviado para activar el flujo")                    

class Inclusion_fallos(inclusion_mac):
     def __init__(self):
        super().__init__()         
     def borrado_fallos(self, tabla, N_hoja):
        self.sharepoint(tabla, N_hoja, "", "descargar")
        wb = load_workbook(tabla)
        ws = wb[N_hoja]
        max_row = ws.max_row
        
        for row in range(ws.max_row, 1, -1):
            valores = [cell.value for cell in ws[row]]
            # Si TODA la fila está vacía
            if all(v is None or str(v).strip() == "" for v in valores):
                ws.delete_rows(row, 1)
        
        # Borrar todas las filas excepto la primera (encabezado)
        if max_row > 1:
            ws.delete_rows(2, max_row - 1)
        
        wb.save(tabla)
        wb.close()
        
        self.sharepoint(tabla, N_hoja, "", "subir")
        #print("✅ Filas eliminadas, se conserva la primera fila como encabezado")   
     def main(self):
        
        print("Recorriendo Fallos")
        # Las siguientes líneas están comentadas en tu código original
        Tabla_Aprobacion2 = self.sharepoint("Aprobados.xlsx", "Sheet2", " ", "descargar")
        Tabla_Aprobacion = Tabla_Aprobacion2[Tabla_Aprobacion2["ESTADO"].isin(["ERROR", "PENDIENTE"])]
        
        wb = load_workbook("Aprobados.xlsx")
        ws = wb["Sheet2"]
        # 1. Encontrar el índice de la columna "ESTADO"
        headers = [cell.value for cell in ws[1]]
        col_estado = headers.index("ESTADO") + 1  # openpyxl es 1-based

        # 2. Recorrer de abajo hacia arriba
        for row in range(ws.max_row, 1, -1):
            valor = ws.cell(row=row, column=col_estado).value
            if valor and str(valor).strip().upper() == "ERROR":
                ws.delete_rows(row)
        wb.save("Aprobados.xlsx")
    
        print(Tabla_Aprobacion.head(10))
        self.recorrer(Tabla_Aprobacion)
        
        print("MACS FALLAS INCLUIDAS")
# Para ejecutar directamente este archivo
if __name__ == "__main__":
    proceso = inclusion_mac()
    proceso.main()
